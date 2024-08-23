# -*- coding: UTF-8 -*-
#import zhuyincheck as chk4char #wrong way to import py file will got TypeError: 'module' object is not callable
#from zhuyincheck import zhuyincheck
#import zhuyincheck
	
#check zhuyin
# ␢ is blank character with code: U+2422
def chk4char(zhuyin_input):
#    global zhuyin_output
    consonant = ["ㄅ","ㄆ","ㄇ","ㄈ","ㄉ","ㄊ","ㄋ","ㄌ","ㄍ","ㄎ","ㄏ","ㄐ","ㄑ","ㄒ","ㄓ","ㄔ","ㄕ","ㄖ","ㄗ","ㄘ","ㄙ"] #聲母
    medial = ["ㄧ","ㄨ","ㄩ"]#介音
    rhyme = ["ㄚ","ㄛ","ㄜ","ㄝ","ㄞ","ㄟ","ㄠ","ㄡ","ㄢ","ㄣ","ㄤ","ㄥ","ㄦ"]  #韻母 
    tone = ["ˊ","ˇ","ˋ","˙"] #聲調 ('-','/','v','\','.')
    zhuyin_input = zhuyin_input.replace(" ","") #Remove spaces from string
    zhuyin_input_list = list(zhuyin_input) # convert string to list
    tone1=["ˉ"]
#    print(len(zhuyin_input_list))
    if len(zhuyin_input_list) == 4:
        return zhuyin_input        

    if not any([item in zhuyin_input_list for item in tone]): # check if any elements within tone list on zhuyin list?
        zhuyin_input_list = zhuyin_input_list + tone1
#        return zhuyin_input_list
        if len(zhuyin_input_list) == 4:
           zhuyin_input = "".join(zhuyin_input_list) # convert list to string
           return zhuyin_input
    if len(zhuyin_input_list) == 3:
        if not any([item in zhuyin_input_list for item in consonant]):
            zhuyin_input_list.insert(0,"␢")
        elif not any([item in zhuyin_input_list for item in medial]):
            zhuyin_input_list.insert(1,"␢")
        elif not any([item in zhuyin_input_list for item in rhyme]):
            zhuyin_input_list.insert(2,"␢")

    if len(zhuyin_input_list) == 2:
        if not any([item in zhuyin_input_list for item in consonant]) and not any([item in zhuyin_input_list for item in rhyme]):
            zhuyin_input_list.insert(0,"␢")
            zhuyin_input_list.insert(2,"␢")
        elif not any([item in zhuyin_input_list for item in rhyme]) and not any([item in zhuyin_input_list for item in medial]):
            zhuyin_input_list.insert(1,"␢")
            zhuyin_input_list.insert(2,"␢")
        elif not any([item in zhuyin_input_list for item in medial]) and not any([item in zhuyin_input_list for item in consonant]):
            zhuyin_input_list.insert(0,"␢␢")

    zhuyin_input = "".join(zhuyin_input_list) # convert list to string
    return zhuyin_input


from openpyxl import load_workbook as lw
#from openpyxl.utils import get_column_letter

wb = lw('unicode_chinese_zhuyin_dict.xlsx')
ws = wb['Sheet1']

#for i in ws.iter_rows(max_row=0):
#    print(len(i))
#    break

column=ws['B'] #output tuple-->(B1,B2,B3........Bn)
#print(len(column)) # output length   
    
#load excel file
#workbook = load_workbook(filename="unicode_chinese_zhuyin_dict.xlsx")
 
#open workbook
#sheet = workbook.active

#for col in range(1, sheet.max_column + 1):
#    col_letter = get_column_letter(col)
#    max_col_row = len([cell for cell in ws[col_letter] if cell.value])
#    print("Column: {}, Row numbers: {}".format(col_letter, max_col_row)
j=0
for j in range(1, len(column)+1):
#    cellid="B{}".format(j)
#    print(cellid)
    zhuyin=ws["B{}".format(j)].value
#    print(zhuyin)
#    print(zhuyin[3])
    result=chk4char(zhuyin)
    print(result)       
    ws["C{}".format(j)].value = result #Write result to Excel file

wb.save("2unicode_chinese_zhuyin_dict.xlsx")
#import pandas as pd
#file_location = "unicode_chinese_zhuyin_dict.xlsx"
#sheet = pd.read_excel(file_location)
#print(sheet['zhuyin'])
#zhuyin1=sheet['zhuyin']
